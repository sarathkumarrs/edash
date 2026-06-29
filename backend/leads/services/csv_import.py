"""CSV import for leads (Meta-ads friendly).

Two-phase, mirroring contacts.services.csv_import: ``parse_and_validate`` reads
+ validates without writing; ``commit_rows`` writes inside a transaction. Both
re-run validation so the commit endpoint is safe if called directly.

Designed so a Meta Lead Ads export works with little/no editing:
  * `full_name` is accepted and split into first/last (Meta exports a single
    name column). Separate `first_name`/`last_name` also work.
  * `phone_number`, `company_name`, `notes`, `linkedin`, `job title` are
    accepted as aliases of the canonical headers.
  * No `title` (lead headline) is required — Meta leads rarely have one.

A row must have at least one identifier: a name, an email, or a phone.

Duplicate policy (per org): email is hard-errored (the DB enforces a per-org
case-insensitive unique constraint on Lead.email; we mirror it so users get a
clean row message instead of an IntegrityError). Phone/name are not deduped —
leads are intentionally allowed to repeat people across campaigns.

Assigning a lead (assigned_emails) makes it due today via the assignment
signal, so imported+assigned leads surface on the dashboard immediately.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any, Iterable

from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import URLValidator
from django.db import IntegrityError, transaction
from django.db.models.functions import Lower
from django.utils.text import slugify

from accounts.models import Account
from common.models import Profile, Tags, Teams
from common.utils import COUNTRIES, LEAD_SOURCE, LEAD_STATUS
from common.validators import normalize_phone
from leads.models import Lead


# Canonical headers we understand. Everything is optional; row-level rules
# require at least one identifier (name/email/phone).
KNOWN_HEADERS = (
    "first_name",
    "last_name",
    "full_name",
    "title",
    "email",
    "phone",
    "company",
    "job_title",
    "website",
    "linkedin_url",
    "source",
    "status",
    "rating",
    "description",
    "address_line",
    "city",
    "state",
    "postcode",
    "country",
    "assigned_emails",
    "team_names",
    "tags",
)

# Friendly aliases -> canonical header. Applied during header normalization so
# common Meta/CRM export column names just work.
HEADER_ALIASES = {
    "full name": "full_name",
    "name": "full_name",
    "phone_number": "phone",
    "phone number": "phone",
    "company_name": "company",
    "company name": "company",
    "organization": "company",
    "notes": "description",
    "linkedin": "linkedin_url",
    "linkedin url": "linkedin_url",
    "job title": "job_title",
    "lead_source": "source",
    "lead source": "source",
}

# Curated, ordered template columns + an example row (used for the downloadable
# CSV/Excel templates). A subset of KNOWN_HEADERS in a sensible order.
TEMPLATE_HEADERS = [
    "first_name",
    "last_name",
    "email",
    "phone",
    "company",
    "title",
    "job_title",
    "website",
    "linkedin_url",
    "source",
    "status",
    "rating",
    "description",
    "address_line",
    "city",
    "state",
    "postcode",
    "country",
    "assigned_emails",
    "team_names",
    "tags",
]
TEMPLATE_EXAMPLE = [
    "Jane",
    "Roe",
    "jane.roe@example.com",
    "+1 (202) 555-9911",
    "Acme Inc",
    "Website enquiry",
    "Head of Ops",
    "https://acme.test",
    "https://linkedin.com/in/janeroe",
    "campaign",
    "assigned",
    "HOT",
    "Came from Meta lead ad",
    "500 Main St",
    "Austin",
    "TX",
    "78701",
    "US",
    "agent@yourco.com",
    "Sales",
    "meta;q3-campaign",
]

MAX_ROWS = 5000
NAME_MAX_LEN = 255
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PHONE_RE = re.compile(r"^[\d\s\-\(\)\+\.]{7,25}$")
_url_validator = URLValidator(schemes=("http", "https"))
COUNTRY_CODES = {code.upper() for code, _label in COUNTRIES}
STATUS_VALUES = {value for value, _label in LEAD_STATUS}
SOURCE_VALUES = {value for value, _label in LEAD_SOURCE}
RATING_VALUES = {"HOT", "WARM", "COLD"}
DEFAULT_STATUS = "assigned"


@dataclass
class RowError:
    row: int
    field: str
    message: str

    def to_dict(self) -> dict[str, Any]:
        return {"row": self.row, "field": self.field, "message": self.message}


@dataclass
class ValidatedRow:
    row: int
    first_name: str = ""
    last_name: str = ""
    title: str | None = None
    email: str | None = None
    phone: str | None = None
    company: str | None = None
    job_title: str | None = None
    website: str | None = None
    linkedin_url: str | None = None
    source: str | None = None
    status: str = DEFAULT_STATUS
    rating: str | None = None
    description: str | None = None
    address_line: str | None = None
    city: str | None = None
    state: str | None = None
    postcode: str | None = None
    country: str | None = None
    assigned_ids: list[str] = field(default_factory=list)
    team_ids: list[str] = field(default_factory=list)
    tag_names: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "row": self.row,
            "first_name": self.first_name,
            "last_name": self.last_name,
            "email": self.email,
            "phone": self.phone,
            "company": self.company,
            "status": self.status,
        }


@dataclass
class SkippedRow:
    """A row skipped as a duplicate (only when skip_duplicates is on)."""

    row: int
    reason: str

    def to_dict(self) -> dict[str, Any]:
        return {"row": self.row, "reason": self.reason}


@dataclass
class ImportResult:
    valid: list[ValidatedRow]
    errors: list[RowError]
    header_error: str | None = None
    skipped: list[SkippedRow] = field(default_factory=list)

    @property
    def summary(self) -> dict[str, int]:
        return {
            "total": len(self.valid)
            + len({e.row for e in self.errors})
            + len(self.skipped),
            "valid": len(self.valid),
            "invalid": len({e.row for e in self.errors}),
            "skipped": len(self.skipped),
        }

    def to_dict(self) -> dict[str, Any]:
        return {
            "header_error": self.header_error,
            "valid": [r.to_dict() for r in self.valid],
            "errors": [e.to_dict() for e in self.errors],
            "skipped": [s.to_dict() for s in self.skipped],
            "summary": self.summary,
        }


@dataclass
class _RefMaps:
    accounts: dict[str, str]
    profiles: dict[str, str]
    teams: dict[str, str]
    existing_emails: set[str]
    existing_phones: set[str]


def _decode(file_bytes: bytes) -> str | None:
    for encoding in ("utf-8-sig", "utf-8"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return None


def _cell_to_str(value: Any) -> str:
    """Stringify an Excel cell the way a CSV would hold it."""
    from datetime import date, datetime

    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        # Excel stores plain integers (e.g. a phone) as floats; avoid "1.0".
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _rows_from_bytes(file_bytes: bytes) -> tuple[list[list[str]] | None, str | None]:
    """Return (rows, error). Accepts .xlsx (zip magic) or UTF-8 CSV.

    rows is a list of rows, each a list of stringified cell values, matching
    what csv.reader produces — so the rest of the pipeline is format-agnostic.
    """
    # XLSX is a zip archive; sniff the magic bytes rather than trusting the
    # filename so a mislabeled file fails cleanly.
    if file_bytes[:4] == b"PK\x03\x04":
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows = [
                [_cell_to_str(cell) for cell in row]
                for row in ws.iter_rows(values_only=True)
            ]
            wb.close()
            return rows, None
        except Exception:
            return None, "Could not read the Excel file. Re-save it as .xlsx and try again."

    text = _decode(file_bytes)
    if text is None:
        return None, (
            "File could not be read. Upload a .csv (UTF-8) or .xlsx file."
        )
    return list(csv.reader(io.StringIO(text))), None


def _normalize_headers(raw_headers: Iterable[str]) -> list[str]:
    out = []
    for h in raw_headers:
        key = (h or "").strip().lower()
        out.append(HEADER_ALIASES.get(key, key))
    return out


def _split_multi(raw: str) -> list[str]:
    if not raw:
        return []
    return [part.strip() for part in raw.split(";") if part.strip()]


def _split_full_name(full_name: str) -> tuple[str, str]:
    parts = full_name.strip().split(None, 1)
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def parse_and_validate(
    file_bytes: bytes, org, skip_duplicates: bool = False
) -> ImportResult:
    rows, read_error = _rows_from_bytes(file_bytes)
    if read_error:
        return ImportResult(valid=[], errors=[], header_error=read_error)
    if not rows:
        return ImportResult(valid=[], errors=[], header_error="File is empty")

    headers = _normalize_headers(rows[0])
    unknown = [h for h in headers if h and h not in KNOWN_HEADERS]
    if unknown:
        return ImportResult(
            valid=[],
            errors=[],
            header_error=f"Unknown column(s): {', '.join(unknown)}. Rename or remove them, or download the template.",
        )

    data_rows = rows[1:]
    if len(data_rows) > MAX_ROWS:
        return ImportResult(
            valid=[],
            errors=[],
            header_error=f"Too many rows ({len(data_rows)}); limit is {MAX_ROWS}",
        )

    parsed: list[tuple[int, dict[str, str]]] = []
    for idx, raw_row in enumerate(data_rows, start=1):
        if not any((cell or "").strip() for cell in raw_row):
            continue
        record = {
            h: (raw_row[i].strip() if i < len(raw_row) else "")
            for i, h in enumerate(headers)
            if h
        }
        parsed.append((idx, record))

    refs = _build_ref_maps(parsed, org)

    valid: list[ValidatedRow] = []
    errors: list[RowError] = []
    skipped: list[SkippedRow] = []
    seen_emails: dict[str, int] = {}  # email.lower() -> row number
    seen_phones: dict[str, int] = {}  # normalized phone -> row number

    for idx, record in parsed:
        # Field-level validation first (format, required, references).
        row_errors, validated = _validate_and_build(idx, record, refs)
        if row_errors:
            errors.extend(row_errors)
            continue

        # Duplicate detection on email + phone, against earlier rows in the
        # file and existing leads in the org.
        dup_field, dup_msg = _duplicate_reason(
            validated, refs, seen_emails, seen_phones
        )
        if dup_field:
            if skip_duplicates:
                skipped.append(SkippedRow(idx, dup_msg))
            else:
                errors.append(RowError(idx, dup_field, dup_msg))
            continue

        valid.append(validated)
        if validated.email:
            seen_emails[validated.email.lower()] = idx
        normalized = normalize_phone(validated.phone) if validated.phone else None
        if normalized:
            seen_phones[normalized] = idx

    return ImportResult(valid=valid, errors=errors, skipped=skipped)


def _duplicate_reason(
    validated: ValidatedRow,
    refs: _RefMaps,
    seen_emails: dict[str, int],
    seen_phones: dict[str, int],
) -> tuple[str | None, str]:
    """Return (field, message) if the row duplicates an email or phone, else (None, '')."""
    if validated.email:
        key = validated.email.lower()
        prior = seen_emails.get(key)
        if prior:
            return "email", f"Duplicate email also used by row {prior} in this file"
        if key in refs.existing_emails:
            return "email", "A lead with this email already exists in your organization"
    if validated.phone:
        normalized = normalize_phone(validated.phone)
        if normalized:
            prior = seen_phones.get(normalized)
            if prior:
                return "phone", f"Duplicate phone also used by row {prior} in this file"
            if normalized in refs.existing_phones:
                return "phone", "A lead with this phone number already exists in your organization"
    return None, ""


def _build_ref_maps(parsed: list[tuple[int, dict[str, str]]], org) -> _RefMaps:
    company_names: set[str] = set()
    assigned_emails: set[str] = set()
    team_names: set[str] = set()
    candidate_emails: set[str] = set()
    candidate_phones: set[str] = set()

    for _idx, record in parsed:
        company = record.get("company", "")
        if company:
            company_names.add(company.lower())
        for email in _split_multi(record.get("assigned_emails", "")):
            assigned_emails.add(email.lower())
        for team in _split_multi(record.get("team_names", "")):
            team_names.add(team.lower())
        email = record.get("email", "")
        if email:
            candidate_emails.add(email.lower())
        phone = record.get("phone", "")
        if phone:
            normalized = normalize_phone(phone)
            if normalized:
                candidate_phones.add(normalized)

    accounts: dict[str, str] = {}
    if company_names:
        for pk, name_lower in (
            Account.objects.filter(org=org)
            .annotate(name_lower=Lower("name"))
            .filter(name_lower__in=company_names)
            .values_list("id", "name_lower")
        ):
            accounts.setdefault(name_lower, str(pk))

    profiles: dict[str, str] = {}
    if assigned_emails:
        for pk, email_lower in (
            Profile.objects.filter(org=org, is_active=True)
            .annotate(email_lower=Lower("user__email"))
            .filter(email_lower__in=assigned_emails)
            .values_list("id", "email_lower")
        ):
            profiles.setdefault(email_lower, str(pk))

    teams: dict[str, str] = {}
    if team_names:
        for pk, name_lower in (
            Teams.objects.filter(org=org)
            .annotate(name_lower=Lower("name"))
            .filter(name_lower__in=team_names)
            .values_list("id", "name_lower")
        ):
            teams.setdefault(name_lower, str(pk))

    existing_emails: set[str] = set()
    if candidate_emails:
        existing_emails = set(
            Lead.objects.filter(org=org)
            .exclude(email__isnull=True)
            .exclude(email="")
            .annotate(email_lower=Lower("email"))
            .filter(email_lower__in=candidate_emails)
            .values_list("email_lower", flat=True)
        )

    existing_phones: set[str] = set()
    if candidate_phones:
        # Phone has no DB constraint / normalized column, so scan this org's
        # leads with a phone and normalize in Python. Bounded by org size.
        for raw_phone in (
            Lead.objects.filter(org=org)
            .exclude(phone__isnull=True)
            .exclude(phone="")
            .values_list("phone", flat=True)
        ):
            normalized = normalize_phone(raw_phone)
            if normalized and normalized in candidate_phones:
                existing_phones.add(normalized)

    return _RefMaps(
        accounts=accounts,
        profiles=profiles,
        teams=teams,
        existing_emails=existing_emails,
        existing_phones=existing_phones,
    )


def _validate_and_build(
    idx: int, record: dict, refs: _RefMaps
) -> tuple[list[RowError], ValidatedRow | None]:
    """Field-level validation only. Duplicate detection happens separately in
    parse_and_validate so it can be routed to errors or skipped rows."""
    errors: list[RowError] = []

    first_name = record.get("first_name", "")
    last_name = record.get("last_name", "")
    full_name = record.get("full_name", "")
    if full_name and not first_name and not last_name:
        first_name, last_name = _split_full_name(full_name)

    for fld, val in (("first_name", first_name), ("last_name", last_name)):
        if val and len(val) > NAME_MAX_LEN:
            errors.append(RowError(idx, fld, f"Exceeds {NAME_MAX_LEN} characters"))

    email = record.get("email", "")
    if email and not EMAIL_RE.match(email):
        errors.append(RowError(idx, "email", f"'{email}' is not a valid email"))

    phone = record.get("phone", "")
    if phone and not PHONE_RE.match(phone):
        errors.append(
            RowError(idx, "phone", "Phone must be 7-25 characters of digits and separators (+ - ( ) . space)")
        )

    # A lead needs something to identify it.
    if not first_name and not last_name and not email and not phone:
        errors.append(RowError(idx, "row", "Row needs at least a name, email, or phone"))

    website = record.get("website", "")
    if website:
        try:
            _url_validator(website)
        except DjangoValidationError:
            errors.append(RowError(idx, "website", "Must be a valid URL (http:// or https://)"))

    linkedin_url = record.get("linkedin_url", "")
    if linkedin_url:
        try:
            _url_validator(linkedin_url)
        except DjangoValidationError:
            errors.append(RowError(idx, "linkedin_url", "Must be a valid URL (http:// or https://)"))

    status_raw = record.get("status", "")
    status = status_raw.strip().lower() if status_raw else DEFAULT_STATUS
    if status not in STATUS_VALUES:
        errors.append(
            RowError(idx, "status", f"'{status_raw}' is not a valid status ({', '.join(sorted(STATUS_VALUES))})")
        )

    source_raw = record.get("source", "")
    source = source_raw.strip().lower() or None
    if source and source not in SOURCE_VALUES:
        # Source is a free-text-with-choices field; accept unknown values
        # rather than block a Meta export, but keep it lowercased.
        pass

    rating_raw = record.get("rating", "")
    rating = rating_raw.strip().upper() or None
    if rating and rating not in RATING_VALUES:
        errors.append(RowError(idx, "rating", "Rating must be Hot, Warm, or Cold"))
        rating = None

    country_raw = record.get("country", "")
    country: str | None = None
    if country_raw:
        country = country_raw.strip().upper()
        if country not in COUNTRY_CODES:
            errors.append(RowError(idx, "country", f"'{country_raw}' is not a known country code"))
            country = None

    for length_field in ("title", "company", "job_title", "address_line", "city", "state"):
        val = record.get(length_field, "")
        if val and len(val) > NAME_MAX_LEN:
            errors.append(RowError(idx, length_field, f"Exceeds {NAME_MAX_LEN} characters"))

    postcode = record.get("postcode", "")
    if postcode and len(postcode) > 64:
        errors.append(RowError(idx, "postcode", "Exceeds 64 characters"))

    assigned_ids: list[str] = []
    for assigned_email in _split_multi(record.get("assigned_emails", "")):
        if not EMAIL_RE.match(assigned_email):
            errors.append(RowError(idx, "assigned_emails", f"'{assigned_email}' is not a valid email"))
            continue
        resolved = refs.profiles.get(assigned_email.lower())
        if resolved is None:
            errors.append(RowError(idx, "assigned_emails", f"No active member with email '{assigned_email}'"))
        else:
            assigned_ids.append(resolved)

    team_ids: list[str] = []
    for team_name in _split_multi(record.get("team_names", "")):
        resolved = refs.teams.get(team_name.lower())
        if resolved is None:
            errors.append(RowError(idx, "team_names", f"No team named '{team_name}'"))
        else:
            team_ids.append(resolved)

    if errors:
        return errors, None

    return [], ValidatedRow(
        row=idx,
        first_name=first_name,
        last_name=last_name,
        title=record.get("title") or None,
        email=email or None,
        phone=phone or None,
        company=record.get("company") or None,
        job_title=record.get("job_title") or None,
        website=website or None,
        linkedin_url=linkedin_url or None,
        source=source,
        status=status,
        rating=rating,
        description=record.get("description") or None,
        address_line=record.get("address_line") or None,
        city=record.get("city") or None,
        state=record.get("state") or None,
        postcode=postcode or None,
        country=country,
        assigned_ids=assigned_ids,
        team_ids=team_ids,
        tag_names=_split_multi(record.get("tags", "")),
    )


def commit_rows(file_bytes: bytes, org, profile, skip_duplicates: bool = False) -> dict[str, Any]:
    result = parse_and_validate(file_bytes, org, skip_duplicates=skip_duplicates)
    if result.header_error:
        return {"error": True, "header_error": result.header_error, "created": 0}
    if result.errors:
        return {
            "error": True,
            "message": "Fix the invalid rows before importing",
            "errors": [e.to_dict() for e in result.errors],
            "created": 0,
        }
    try:
        out = _commit_validated(result.valid, org, profile)
        out["skipped"] = len(result.skipped)
        return out
    except IntegrityError as exc:
        return {
            "error": True,
            "message": (
                "A lead was created concurrently that conflicts with this import "
                "(likely a duplicate email). Re-run preview and try again."
            ),
            "detail": str(exc),
            "created": 0,
        }


@transaction.atomic
def _commit_validated(rows: list[ValidatedRow], org, profile) -> dict[str, Any]:
    created_ids: list[str] = []
    tag_cache: dict[str, Tags] = {}

    for vr in rows:
        lead = Lead.objects.create(
            first_name=vr.first_name or None,
            last_name=vr.last_name or None,
            title=vr.title,
            email=vr.email,
            phone=vr.phone,
            company_name=vr.company,
            job_title=vr.job_title,
            website=vr.website,
            linkedin_url=vr.linkedin_url,
            source=vr.source,
            status=vr.status,
            rating=vr.rating,
            description=vr.description,
            address_line=vr.address_line,
            city=vr.city,
            state=vr.state,
            postcode=vr.postcode,
            country=vr.country,
            org=org,
            created_by=profile.user,
        )
        if vr.assigned_ids:
            # Triggers the assignment signal -> next_follow_up = today.
            lead.assigned_to.set(vr.assigned_ids)
        if vr.team_ids:
            lead.teams.set(vr.team_ids)
        if vr.tag_names:
            tag_objs = [_get_or_create_tag(name, org, tag_cache) for name in vr.tag_names]
            lead.tags.set(tag_objs)
        created_ids.append(str(lead.id))

    return {"error": False, "created": len(created_ids), "ids": created_ids}


def _get_or_create_tag(name: str, org, cache: dict[str, Tags]) -> Tags:
    slug = slugify(name) or name.lower()
    if slug in cache:
        return cache[slug]
    tag, _ = Tags.objects.get_or_create(slug=slug, org=org, defaults={"name": name})
    cache[slug] = tag
    return tag
